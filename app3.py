import subprocess
from datetime import datetime
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.chrome.service import Service
import openpyxl
from openpyxl.styles import Font

# Configurações do banco de dados
DB_PATH = 'C:\\robo\\CONTROLE.FDB'
DB_USER = 'sysdba'
DB_PASSWORD = 'Q5QIST'
ISQL_PATH = 'C:\\Program Files\\Firebird\\Firebird_2_5\\bin\\isql.exe'
SCRIPT_SQL_PATH = 'C:\\robo\\script.sql'  # Caminho correto para o script SQL

# Configurações de logging
logging.basicConfig(
    filename='script.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
)

# Credenciais Shipsgo
SHIPS_GO_USER = 'ivan.moraes@aduantech.com.br'
SHIPS_GO_PASSWORD = '.v8tX8nTTuNFe6E'
SHIPS_GO_URL = 'https://shipsgo.com/pt'

# Função para verificar se o Firebird está rodando
def check_firebird_status():
    try:
        result = subprocess.run(['tasklist'], capture_output=True, text=True)
        if 'fbserver.exe' in result.stdout or 'fb_inet_server.exe' in result.stdout:
            logging.info("Firebird está rodando.")
            print("Firebird está rodando.")
            return True
        else:
            logging.warning("Firebird não está rodando.")
            print("Firebird não está rodando.")
            return False
    except Exception as e:
        logging.error(f"Erro ao verificar o status do Firebird: {e}")
        print(f"Erro ao verificar o status do Firebird: {e}")
        return False

# Função para executar a consulta SQL usando isql
def execute_sql_query():
    if not check_firebird_status():
        logging.error("Execução abortada: Firebird não está rodando.")
        print("Execução abortada: Firebird não está rodando.")
        return None

    try:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        output_file = f'saida_{timestamp}.txt'
        with open(output_file, 'w') as file:
            process = subprocess.run(
                [ISQL_PATH, DB_PATH, '-user', DB_USER, '-password', DB_PASSWORD, '-i', SCRIPT_SQL_PATH],
                stdout=file,
                stderr=subprocess.PIPE,
                text=True
            )
        if process.stderr:
            logging.error(f"Erro do isql: {process.stderr}")
            print(f"Erro do isql: {process.stderr}")
        else:
            logging.info(f"Consulta SQL executada com sucesso. Resultados salvos em {output_file}.")
            print(f"Consulta SQL executada com sucesso. Resultados salvos em {output_file}.")
        return output_file
    except Exception as e:
        logging.error(f"Erro ao executar a consulta SQL: {e}")
        print(f"Erro ao executar a consulta SQL: {e}")
        return None

# Função para validar os dados extraídos
def validate_data(file_path):
    try:
        with open(file_path, 'r') as file:
            content = file.read()
            if "FATURA" in content and "CLIENTE" in content and "CONTEINER" in content:
                logging.info(f"Dados extraídos do arquivo {file_path} estão válidos.")
                print(f"Dados extraídos do arquivo {file_path} estão válidos.")
                return True, content
            else:
                logging.warning(f"Arquivo {file_path} não possui todos os parâmetros necessários.")
                print(f"Arquivo {file_path} não possui todos os parâmetros necessários.")
                return False, None
    except Exception as e:
        logging.error(f"Erro ao validar os dados extraídos: {e}")
        print(f"Erro ao validar os dados extraídos: {e}")
        return False, None

# Função para processar os resultados e validar os dados
def process_results(output_file):
    if output_file:
        valid, content = validate_data(output_file)
        if valid:
            logging.info(f"Arquivo {output_file} encontrado:")
            print(f"Arquivo {output_file} encontrado:")
            return extract_tracking_numbers(content), content
        else:
            logging.warning("Dados extraídos não são válidos.")
            print("Dados extraídos não são válidos.")
            return None, None
    else:
        logging.warning("Nenhum arquivo de saída foi gerado.")
        print("Nenhum arquivo de saída foi gerado.")
        return None, None

# Função para extrair os números de contêineres do conteúdo do arquivo
def extract_tracking_numbers(content):
    lines = content.splitlines()
    tracking_numbers = []
    for line in lines:
        if "CONTEINER" in line or "===============" in line:
            continue
        parts = line.split()
        if len(parts) >= 3:
            container_number = parts[-1].replace("-", "").replace(".", "")
            if container_number and container_number != "<null>":
                tracking_numbers.append(container_number)
    return tracking_numbers

# Função para atualizar a planilha
def atualizar_planilha(container_number, shipment_number, movimentos, dados_txt):
    try:
        # Abrir ou criar a planilha
        try:
            workbook = openpyxl.load_workbook('Shipments.xlsx')
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # Criar uma nova aba para o contêiner
        sheet = workbook.create_sheet(title=f"Container {container_number}")

        # Formatar a primeira linha
        bold_font = Font(bold=True)
        headers = ["Fatura", "Cliente", "Container"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = bold_font

        # Preencher os dados do TXT
        txt_lines = dados_txt.splitlines()
        fatura_cliente_container = txt_lines[3].split()
        sheet.append([fatura_cliente_container[0], " ".join(fatura_cliente_container[1:-1]), fatura_cliente_container[-1]])

        # Formatar a segunda seção
        headers = ["Location", "Moves", "Date", "Vessel", "Check"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col_num, value=header)
            cell.font = bold_font

        # Preencher os dados dos movimentos
        for movimento in movimentos:
            if movimento[4] != "-":
                movimento[4] = "OK"
            sheet.append(movimento)

        # Salvar a planilha
        workbook.save('Shipments.xlsx')
        logging.info(f"Planilha atualizada com sucesso para o contêiner {container_number}.")
        print(f"Planilha atualizada com sucesso para o contêiner {container_number}.")
    except Exception as e:
        logging.error(f"Erro ao atualizar a planilha: {e}")
        print(f"Erro ao atualizar a planilha: {e}")

# Função para realizar o login no Shipsgo
def login_shipsgo(driver, tracking_numbers, dados_txt):
    driver.get(SHIPS_GO_URL)
    wait = WebDriverWait(driver, 10)
    login_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="navbarSupportedContent"]/div/a[1]')))
    login_button.click()

    time.sleep(2.5)  # Pausa de 2.5 segundos

    email_input = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="input-email"]')))
    password_input = driver.find_element(By.XPATH, '//*[@id="input-password"]')

    email_input.send_keys(SHIPS_GO_USER)
    password_input.send_keys(SHIPS_GO_PASSWORD)
    password_input.send_keys(Keys.RETURN)

    # Verifique se o login foi bem-sucedido
    wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="side-menu"]/li[7]/a/span')))
    logging.info("Login bem-sucedido no Shipsgo.")
    print("Login bem-sucedido no Shipsgo.")

    # Navegar para o menu "My Shipments"
    my_shipment_button = driver.find_element(By.XPATH, '//*[@id="side-menu"]/li[7]/a/span')
    my_shipment_button.click()

    time.sleep(2.5)  # Pausa de 2.5 segundos

    # Clicar no botão "Add Filter"
    add_filter_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="table-parent"]/div[1]/button[2]')))
    add_filter_button.click()

    time.sleep(2.5)  # Pausa de 2.5 segundos

    # Clicar no seletor de filtro
    filter_selector = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="table-parent"]/div[3]/div/div/div[2]/form/div/div[1]/span/span[1]/span/span[2]')))
    filter_selector.click()

    time.sleep(2.5)  # Pausa de 2.5 segundos

    # Simular as teclas de seta para baixo 3 vezes e pressionar Enter
    actions = webdriver.ActionChains(driver)
    actions.send_keys(Keys.DOWN).perform()
    time.sleep(0.5)
    actions.send_keys(Keys.DOWN).perform()
    time.sleep(0.5)
    actions.send_keys(Keys.DOWN).perform()
    time.sleep(0.5)
    actions.send_keys(Keys.RETURN).perform()

    time.sleep(2.5)  # Pausa de 2.5 segundos

    # Função para inserir o número de contêiner e processar
    def inserir_container(number):
        input_field_xpath = '//*[@id="table-parent"]/div[3]/div/div/div[2]/form/div/div[3]/input'
        input_field = wait.until(EC.element_to_be_clickable((By.XPATH, input_field_xpath)))
        input_field.click()

        try:
            print(f"Inserindo número de contêiner: {number}")  # Log para depuração
            input_field.clear()
            input_field.send_keys(number)
            input_field.send_keys(Keys.RETURN)
            time.sleep(2.5)  # Pausa de 2.5 segundos para observar a ação
        except Exception as e:
            logging.error(f"Erro ao inserir o número do contêiner: {e}")
            print(f"Erro ao inserir o número do contêiner: {e}")

        # Verificar se o contêiner já está registrado
        try:
            no_data_message = driver.find_element(By.XPATH, '//*[contains(text(),"There is nothing to see here!")]')
            # Se a mensagem de "nothing to see here" estiver presente, registrar o contêiner
            logging.info("Contêiner não registrado, procedendo com o registro.")
            print("Contêiner não registrado, procedendo com o registro.")

            # Clicar no botão para adicionar novo contêiner
            add_new_shipment_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="table-parent"]/div[1]/a')))
            add_new_shipment_button.click()

            # Clicar no campo de entrada do contêiner
            container_input_xpath = '//*[@id="input-container-number"]'
            container_input = wait.until(EC.element_to_be_clickable((By.XPATH, container_input_xpath)))
            container_input.click()
            container_input.send_keys(number)

            # Clicar no botão para criar o envio
            create_shipment_button_xpath = '//*[@id="trackSingleForm"]/div[10]/button'
            create_shipment_button = wait.until(EC.element_to_be_clickable((By.XPATH, create_shipment_button_xpath)))
            create_shipment_button.click()

            # Capturar o número do novo envio
            success_message_xpath = '/html/body/div[3]/div/div/div[2]'
            success_message = wait.until(EC.presence_of_element_located((By.XPATH, success_message_xpath))).text
            shipment_number = success_message.split('#')[1].split(')')[0]

            # Atualizar a planilha com o novo envio
            movimentos = []  # Obter movimentos como feito anteriormente
            atualizar_planilha(number, shipment_number, movimentos, dados_txt)

        except Exception:
            # Caso o contêiner já esteja registrado
            logging.info(f"Contêiner já registrado com envio: {number}")
            print(f"Contêiner já registrado com envio: {number}")

            # Clicar no botão para ver detalhes do envio
            detalhes_button_xpath = '//*[@id="table-parent"]/div[2]/table/tbody/tr/td[1]/button'
            detalhes_button = wait.until(EC.element_to_be_clickable((By.XPATH, detalhes_button_xpath)))
            detalhes_button.click()

            # Capturar os movimentos do envio
            movimentos = []
            for i in range(1, 7):  # Ajuste o range conforme necessário
                location_xpath = f'//*[@id="movements"]/div/table/tbody/tr[{i}]/td[1]'
                moves_xpath = f'//*[@id="movements"]/div/table/tbody/tr[{i}]/td[2]'
                date_xpath = f'//*[@id="movements"]/div/table/tbody/tr[{i}]/td[3]'
                vessel_xpath = f'//*[@id="movements"]/div/table/tbody/tr[{i}]/td[4]'
                check_xpath = f'//*[@id="movements"]/div/table/tbody/tr[{i}]/td[5]'

                try:
                    location = wait.until(EC.presence_of_element_located((By.XPATH, location_xpath))).text
                    moves = wait.until(EC.presence_of_element_located((By.XPATH, moves_xpath))).text
                    date = wait.until(EC.presence_of_element_located((By.XPATH, date_xpath))).text
                    vessel = wait.until(EC.presence_of_element_located((By.XPATH, vessel_xpath))).text
                    check = wait.until(EC.presence_of_element_located((By.XPATH, check_xpath))).text
                    if check != "-":
                        check = "OK"
                    movimentos.append([location, moves, date, vessel, check])
                except Exception as ex:
                    logging.error(f"Erro ao capturar dados de movimento: {ex}")
                    print(f"Erro ao capturar dados de movimento: {ex}")

            # Atualizar a planilha com os movimentos do contêiner existente
            shipment_number = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="table-parent"]/div[2]/table/tbody/tr/td[6]'))).text
            atualizar_planilha(number, shipment_number, movimentos, dados_txt)

        # Fechar o modal com o XPath fornecido
        final_click_xpath = '/html/body/div[3]/div/div/div[1]/button/span'
        try:
            final_click_button = wait.until(EC.element_to_be_clickable((By.XPATH, final_click_xpath)))
            final_click_button.click()
            logging.info("Modal fechado com sucesso.")
            print("Modal fechado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao fechar o modal: {e}")
            print(f"Erro ao fechar o modal: {e}")

        # Clicar no botão "Show Filters"
        show_filters_xpath = '//*[@id="table-parent"]/div[1]/button[2]'
        try:
            show_filters_button = wait.until(EC.element_to_be_clickable((By.XPATH, show_filters_xpath)))
            show_filters_button.click()
            logging.info("Botão 'Show Filters' clicado com sucesso.")
            print("Botão 'Show Filters' clicado com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao clicar no botão 'Show Filters': {e}")
            print(f"Erro ao clicar no botão 'Show Filters': {e}")

        # Clicar no "x" para excluir o valor do contêiner no filtro
        clear_filter_xpath = '//*[@id="table-parent"]/div[3]/div/div/div[2]/div/div/ul/li/button'
        try:
            clear_filter_button = wait.until(EC.element_to_be_clickable((By.XPATH, clear_filter_xpath)))
            clear_filter_button.click()
            logging.info("Valor do contêiner excluído do filtro com sucesso.")
            print("Valor do contêiner excluído do filtro com sucesso.")
        except Exception as e:
            logging.error(f"Erro ao excluir o valor do contêiner do filtro: {e}")
            print(f"Erro ao excluir o valor do contêiner do filtro: {e}")

    # Processar todos os números de contêiner
    for number in tracking_numbers:
        inserir_container(number)
        time.sleep(2.5)  # Pausa entre os processamentos de contêineres

    # Manter o navegador aberto para inspeção manual
    print("Navegador mantido aberto para inspeção manual.")
    logging.info("Navegador mantido aberto para inspeção manual.")
    input("Pressione Enter para encerrar...")

# Função para executar o script
def execute_script():
    logging.info("Iniciando execução do script...")
    print("Iniciando execução do script...")
    output_file, dados_txt = process_results(execute_sql_query())
    if output_file and dados_txt:
        chrome_service = Service(executable_path='C:\\Users\\Administrador\\Documents\\chromedriver-win64\\chromedriver.exe')  # Substitua pelo caminho correto do ChromeDriver
        driver = webdriver.Chrome(service=chrome_service)
        try:
            login_shipsgo(driver, output_file, dados_txt)
        finally:
            driver.quit()
    logging.info("Execução do script finalizada.")
    print("Execução do script finalizada.")

# Executar o script manualmente
if __name__ == "__main__":
    logging.info("Executando o script manualmente...")
    print("Executando o script manualmente...")
    execute_script()
